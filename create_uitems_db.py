import sqlite3
import openpyxl

def create_database_from_excel(excel_path, db_path):
    # Excel dosyasını yükle
    workbook = openpyxl.load_workbook(excel_path)
    
    # SQLite veritabanını oluştur veya aç
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Excel dosyasındaki her bir sayfa (tablo) için işlemleri yap
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # İlk satırdaki başlıkları (sütun isimlerini) al
        columns = []
        for cell in sheet[1]:  # İlk satırdaki hücreler
            columns.append(cell.value)
        
        # SQL sorgusunu tablo yapısına göre dinamik olarak oluştur
        create_table_query = f"CREATE TABLE IF NOT EXISTS {sheet_name} ("
        for column in columns:
            create_table_query += f"{column} TEXT, "
        create_table_query = create_table_query.rstrip(', ') + ")"  # Fazladan virgülü kaldır ve sorguyu kapat
        cursor.execute(create_table_query)
        
        # Her satırı (ilk satır hariç) veritabanına ekle
        for row in sheet.iter_rows(min_row=2, values_only=True):
            placeholders = ', '.join(['?' for _ in columns])
            insert_query = f"INSERT INTO {sheet_name} ({', '.join(columns)}) VALUES ({placeholders})"
            cursor.execute(insert_query, row)
    
    # Değişiklikleri kaydet ve bağlantıyı kapat
    conn.commit()
    conn.close()
    print("Data has been successfully transferred from Excel to the SQLite database.")

# Kullanım
excel_path = r"C:\Users\Volkan\Documents\unit_1page_b\walldb.xlsx"
db_path = "wall.db"
create_database_from_excel(excel_path, db_path)
